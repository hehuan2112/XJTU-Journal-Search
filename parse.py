# -*- coding: utf8 -*-

import os
import json
import codecs
from openpyxl import load_workbook

files = os.listdir('./data/')

data = {'cates': [], 'journals': []}

for fn in files:

    if not fn.endswith('.xlsx'): continue
    if fn.startswith('~'): continue

    print("* Parsing %s" % fn)
    # Use filename as category and level
    cate, level = fn.split('.')[0].split('_')
    
    # Load sheets
    wb = load_workbook(fn)
    fn = fn.decode('gbk')

    if cate not in data['cates']: data['cates'].append(cate)
    wsnames = wb.get_sheet_names()

    for wsn in wsnames:
        print("* Loading Sheet %s" % wsn)
        ws = wb.get_sheet_by_name(wsn)
        rowNum = 0

        while True:
            rowNum += 1

            try:
                cellv = ws['A%d' % rowNum].value
                if cellv is None: break

                # exception may occur in this line
                jnum = int(cellv)
                jtit = ws['B%d' % rowNum].value
                if cate == 'JJ' or cate == 'RW':
                    jnat = '-'
                    jssn = ws['C%d' % rowNum].value
                else:
                    jnat = ws['C%d' % rowNum].value
                    jssn = ws['D%d' % rowNum].value

                # add to dict
                data['journals'].append([jnum, jtit, jnat, jssn, cate, level])

            except:
                continue
        
        print("* Loaded %s rows" % rowNum)

json.dump(data, codecs.open('journals.json', 'w'))
print('* Done!')